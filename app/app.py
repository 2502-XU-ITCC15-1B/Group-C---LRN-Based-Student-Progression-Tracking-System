from flask import Flask, flash, redirect, render_template, request, send_file, send_from_directory, session, url_for
from functools import wraps
from io import BytesIO
import mysql.connector
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import os
import pandas as pd
import re
import secrets
import tempfile
import time
from werkzeug.security import check_password_hash, generate_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "change-this-secret-key")


@app.route("/healthz")
def healthz():
    return {"status": "ok"}


@app.after_request
def add_no_cache_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, private, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

SUPPORTED_GRADES = [7, 8, 9, 10]
SCHOOL_YEAR_PATTERN = re.compile(r"^\d{4}-\d{4}$")
LRN_PATTERN = re.compile(r"^\d{12}$")
SCHEMA_READY = False
SUPPORTED_UPLOAD_EXTENSIONS = {".xls", ".xlsx", ".csv"}
PENDING_UPLOAD_TTL_SECONDS = 900


def login_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in to access the system.")
            return redirect(url_for("login", next=request.path))

        return view(*args, **kwargs)

    return wrapped_view


def admin_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in to access the system.")
            return redirect(url_for("login", next=request.path))

        if session.get("role") != "admin":
            flash("You do not have permission to perform that action.")
            return redirect(url_for("dashboard"))

        return view(*args, **kwargs)

    return wrapped_view


@app.context_processor
def inject_current_user():
    return {
        "current_user": session.get("username"),
        "current_role": session.get("role"),
        "status_badge_class": status_badge_class,
    }


EDITABLE_RECORD_STATUSES = ["ENROLLED", "TRANSFER_IN", "PENDING_TRANSFER_IN", "TRANSFER_OUT"]
PER_PAGE = 10


def status_badge_class(status):
    classes = {
        "ENROLLED": "bg-success",
        "TRANSFER_IN": "bg-info text-dark",
        "PENDING_TRANSFER_IN": "bg-warning text-dark",
        "TRANSFER_OUT": "bg-danger",
        "MISSING": "bg-secondary",
        "REPEATED": "bg-warning text-dark",
        "COMPLETED": "bg-primary",
        "DELAYED_COMPLETED": "bg-info text-dark",
        "STRAIGHT_PATH": "bg-success",
        "INCOMPLETE": "bg-secondary",
    }
    return classes.get(status, "bg-dark")


def is_valid_school_year(school_year):
    if not school_year or not SCHOOL_YEAR_PATTERN.match(school_year):
        return False

    start_year, end_year = [int(part) for part in school_year.split("-")]
    return end_year == start_year + 1


def read_lis_upload(file):
    filename = (file.filename or "").strip().lower()
    extension = os.path.splitext(filename)[1]

    if extension not in SUPPORTED_UPLOAD_EXTENSIONS:
        raise ValueError("Unsupported file type. Upload a LIS/SF1 file in .xlsx, .xls, or .csv format.")

    if extension == ".csv":
        try:
            return pd.read_csv(file, header=None, dtype=str, encoding="utf-8-sig")
        except UnicodeDecodeError:
            file.stream.seek(0)
            return pd.read_csv(file, header=None, dtype=str, encoding="latin1")

    return pd.read_excel(file, header=None)


def detect_upload_metadata(df):
    detected = {"school_year": None, "grade_level": None}

    for _, row in df.head(15).iterrows():
        for value in row:
            text = str(value).upper().replace("\n", " ").strip()
            if not text or text == "NAN":
                continue

            if detected["school_year"] is None:
                year_match = re.search(r"\b(\d{4})\s*[-–—]\s*(\d{4})\b", text)
                if year_match:
                    school_year = f"{year_match.group(1)}-{year_match.group(2)}"
                    if is_valid_school_year(school_year):
                        detected["school_year"] = school_year

            if detected["grade_level"] is None:
                grade_match = re.search(r"\b(?:GRADE|GRADE LEVEL|GR\.?)\s*[:\-]?\s*(7|8|9|10)\b", text)
                if grade_match:
                    detected["grade_level"] = int(grade_match.group(1))

            if detected["school_year"] and detected["grade_level"]:
                return detected

    return detected


def get_upload_metadata_mismatches(detected, school_year, grade_level):
    mismatches = []

    if detected.get("school_year") and detected["school_year"] != school_year:
        mismatches.append({
            "field": "School Year",
            "detected": detected["school_year"],
            "selected": school_year,
        })

    if detected.get("grade_level") and detected["grade_level"] != grade_level:
        mismatches.append({
            "field": "Grade Level",
            "detected": f"Grade {detected['grade_level']}",
            "selected": f"Grade {grade_level}",
        })

    return mismatches


def get_pending_upload_dir():
    path = os.path.join(tempfile.gettempdir(), "lrn_upload_confirmations")
    os.makedirs(path, exist_ok=True)
    return path


def save_pending_upload(file):
    extension = os.path.splitext(file.filename or "")[1].lower()
    token = secrets.token_urlsafe(24)
    path = os.path.join(get_pending_upload_dir(), f"{token}{extension}")

    file.stream.seek(0)
    with open(path, "wb") as output:
        output.write(file.stream.read())

    pending_uploads = session.get("pending_uploads", {})
    pending_uploads[token] = {
        "path": path,
        "filename": os.path.basename(file.filename or f"upload{extension}"),
        "created_at": time.time(),
    }
    session["pending_uploads"] = pending_uploads
    return token


def pop_pending_upload(token):
    pending_uploads = session.get("pending_uploads", {})
    pending = pending_uploads.pop(token, None)
    session["pending_uploads"] = pending_uploads

    if not pending:
        return None

    is_expired = time.time() - pending.get("created_at", 0) > PENDING_UPLOAD_TTL_SECONDS
    if is_expired or not os.path.exists(pending.get("path", "")):
        try:
            os.remove(pending.get("path", ""))
        except OSError:
            pass
        return None

    return pending


class StoredUpload:
    def __init__(self, path, filename):
        self.filename = filename
        self.stream = open(path, "rb")

    def read(self, *args, **kwargs):
        return self.stream.read(*args, **kwargs)

    def seek(self, *args, **kwargs):
        return self.stream.seek(*args, **kwargs)

    def __getattr__(self, name):
        return getattr(self.stream, name)

    def close(self):
        self.stream.close()


@app.route("/")
def root():
    if "user_id" in session:
        return redirect(url_for("dashboard"))

    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    ensure_schema()

    if "user_id" in session and request.method == "GET":
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            """
            SELECT id, username, password_hash, role, is_active
            FROM users
            WHERE username = %s
            """,
            (username,),
        )
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if not user or not user["is_active"] or not check_password_hash(user["password_hash"], password):
            flash("Invalid username or password.")
            return render_template("login.html", username=username)

        session.clear()
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["role"] = user["role"]
        return redirect(url_for("dashboard"))

    return render_template("login.html")


@app.route("/logout", methods=["GET", "POST"])
@login_required
def logout():
    if request.method == "POST":
        session.clear()
        flash("Logged out successfully.")
        return redirect(url_for("login"))
    return render_template("logout_confirm.html")


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    ensure_schema()

    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if len(new_password) < 8:
            flash("New password must be at least 8 characters.")
            return render_template("change_password.html")

        if new_password != confirm_password:
            flash("New password and confirmation do not match.")
            return render_template("change_password.html")

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            """
            SELECT id, password_hash
            FROM users
            WHERE id = %s
            """,
            (session["user_id"],),
        )
        user = cursor.fetchone()

        if not user or not check_password_hash(user["password_hash"], current_password):
            cursor.close()
            conn.close()
            flash("Current password is incorrect.")
            return render_template("change_password.html")

        cursor.execute(
            """
            UPDATE users
            SET password_hash = %s
            WHERE id = %s
            """,
            (generate_password_hash(new_password), session["user_id"]),
        )
        conn.commit()
        cursor.close()
        conn.close()
        flash("Password changed successfully.")
        return redirect(url_for("dashboard"))

    return render_template("change_password.html")


@app.route("/dashboard")
@login_required
def dashboard():
    ensure_schema()

    conn = get_db_connection()
    cursor = conn.cursor()
    stats = get_dashboard_stats(cursor)
    recent_changes = get_recent_changes(cursor, limit=5)
    grade_distribution = get_grade_distribution(cursor)
    cursor.close()
    conn.close()

    return render_template(
        "dashboard.html",
        stats=stats,
        recent_changes=recent_changes,
        grade_distribution=grade_distribution,
    )


@app.route("/lis-upload")
@login_required
def lis_upload():
    cancel_upload = request.args.get("cancel_upload", "").strip()
    if cancel_upload:
        pending = pop_pending_upload(cancel_upload)
        if pending:
            try:
                os.remove(pending["path"])
            except OSError:
                pass
        return redirect(url_for("lis_upload"))

    return render_template("lis_upload.html", grades=SUPPORTED_GRADES)


@app.route("/records/delete-batch", methods=["POST"])
@admin_required
def delete_batch_records():
    ensure_schema()

    school_year = request.form.get("school_year", "").strip()
    grade_level = request.form.get("grade_level", "").strip()
    confirmation = request.form.get("confirmation", "").strip()

    if not is_valid_school_year(school_year):
        flash("Use school year format YYYY-YYYY with consecutive years before deleting a batch.")
        return redirect(url_for("lis_upload"))

    if not grade_level.isdigit() or int(grade_level) not in SUPPORTED_GRADES:
        flash("Select a valid Grade 7 to Grade 10 batch to delete.")
        return redirect(url_for("lis_upload"))

    if confirmation != "DELETE":
        flash("Type DELETE to confirm batch deletion.")
        return redirect(url_for("lis_upload"))

    grade_level = int(grade_level)
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT COUNT(*)
        FROM student_records
        WHERE school_year = %s
        AND grade_level = %s
        """,
        (school_year, grade_level),
    )
    record_count = cursor.fetchone()[0] or 0

    if record_count == 0:
        cursor.close()
        conn.close()
        flash(f"No Grade {grade_level} records found for {school_year}.")
        return redirect(url_for("lis_upload"))

    cursor.execute(
        """
        DELETE FROM student_change_logs
        WHERE school_year = %s
        AND grade_level = %s
        """,
        (school_year, grade_level),
    )

    cursor.execute(
        """
        DELETE FROM student_records
        WHERE school_year = %s
        AND grade_level = %s
        """,
        (school_year, grade_level),
    )

    cursor.execute(
        """
        DELETE l
        FROM student_change_logs l
        LEFT JOIN student_records r ON l.lrn = r.lrn
        WHERE r.lrn IS NULL
        """
    )

    cursor.execute(
        """
        DELETE s
        FROM students s
        LEFT JOIN student_records r ON s.lrn = r.lrn
        WHERE r.lrn IS NULL
        """
    )

    conn.commit()
    cursor.close()
    conn.close()

    flash(f"Deleted {record_count} Grade {grade_level} records for {school_year}.")
    return redirect(url_for("lis_upload"))


@app.route("/records/delete-all", methods=["POST"])
@admin_required
def delete_all_records():
    ensure_schema()

    confirmation = request.form.get("confirmation", "").strip()

    if confirmation != "DELETE ALL":
        flash("Type DELETE ALL to confirm overall deletion.")
        return redirect(url_for("lis_upload"))

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT COUNT(*) FROM student_records")
        record_count = cursor.fetchone()[0] or 0

        cursor.execute("SELECT COUNT(*) FROM students")
        student_count = cursor.fetchone()[0] or 0

        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
        cursor.execute("TRUNCATE TABLE student_change_logs")
        cursor.execute("TRUNCATE TABLE student_records")
        cursor.execute("TRUNCATE TABLE students")
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")

        conn.commit()
    except Exception:
        conn.rollback()
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        conn.commit()
        raise
    finally:
        cursor.close()
        conn.close()

    flash(f"Deleted all records: {student_count} students and {record_count} grade records.")
    return redirect(url_for("lis_upload"))


@app.route("/students")
@login_required
def students():
    ensure_schema()

    filters = {
        "q": request.args.get("q", "").strip(),
        "grade": request.args.get("grade", "").strip(),
        "status": request.args.get("status", "").strip(),
        "year": request.args.get("year", "").strip(),
        "page": request.args.get("page", "1").strip(),
    }
    page = int(filters["page"]) if filters["page"].isdigit() and int(filters["page"]) > 0 else 1

    conn = get_db_connection()
    cursor = conn.cursor()
    total_matching = count_student_records(cursor, filters)
    total_pages = max(1, (total_matching + PER_PAGE - 1) // PER_PAGE)
    page = min(page, total_pages)
    filters["page"] = str(page)
    records = get_student_records(cursor, filters, page, PER_PAGE)
    stats = get_dashboard_stats(cursor)
    cursor.close()
    conn.close()

    return render_template(
        "records_page.html",
        records=records,
        grades=SUPPORTED_GRADES,
        statuses=EDITABLE_RECORD_STATUSES,
        filters=filters,
        stats=stats,
        page=page,
        per_page=PER_PAGE,
        total_matching=total_matching,
        total_pages=total_pages,
    )


@app.route("/students/add", methods=["POST"])
@admin_required
def add_student():
    ensure_schema()

    lrn = request.form.get("lrn", "").strip()
    name = request.form.get("name", "").strip()
    gender = normalize_gender(request.form.get("gender", ""))
    school_year = request.form.get("school_year", "").strip()
    grade_level = request.form.get("grade_level", "").strip()
    status = request.form.get("status", "").strip()
    remarks = normalize_remarks(request.form.get("remarks", ""))
    confirmation = request.form.get("confirmation", "").strip()

    if confirmation != "ADD STUDENT":
        flash("Type ADD STUDENT to confirm manual student creation.")
        return redirect(url_for("students"))

    if not LRN_PATTERN.match(lrn):
        flash("Enter a valid 12-digit LRN before adding a student.")
        return redirect(url_for("students"))

    if not name:
        flash("Student name is required.")
        return redirect(url_for("students"))

    if gender not in {"MALE", "FEMALE"}:
        flash("Select a valid sex value.")
        return redirect(url_for("students"))

    if not is_valid_school_year(school_year):
        flash("Use school year format YYYY-YYYY with consecutive years before adding a student.")
        return redirect(url_for("students"))

    if not grade_level.isdigit() or int(grade_level) not in SUPPORTED_GRADES:
        flash("Select a valid Grade 7 to Grade 10 record.")
        return redirect(url_for("students"))

    if status not in EDITABLE_RECORD_STATUSES:
        flash("Select a valid student status.")
        return redirect(url_for("students"))

    grade_level = int(grade_level)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT name
        FROM students
        WHERE lrn = %s
        LIMIT 1
        """,
        (lrn,),
    )
    existing_student = cursor.fetchone()

    if existing_student:
        cursor.close()
        conn.close()
        flash(f"Duplicate LRN rejected. {lrn} already belongs to {existing_student[0]}.")
        return redirect(url_for("students"))

    cursor.execute(
        """
        SELECT id
        FROM student_records
        WHERE lrn = %s
        AND school_year = %s
        LIMIT 1
        """,
        (lrn, school_year),
    )

    if cursor.fetchone():
        cursor.close()
        conn.close()
        flash(f"Duplicate school year rejected. {lrn} already has a record in {school_year}.")
        return redirect(url_for("students"))

    changed_by = session.get("username")
    changes_logged = upsert_student(cursor, lrn, name, gender, school_year, grade_level, changed_by)
    record_action, record_changes = upsert_student_record(
        cursor,
        lrn,
        school_year,
        grade_level,
        gender,
        status,
        remarks,
        changed_by,
    )
    changes_logged += record_changes
    log_change(cursor, lrn, "manual.student_record", "", f"Added Grade {grade_level} record for {school_year}", school_year, grade_level, changed_by)
    conn.commit()
    cursor.close()
    conn.close()

    flash(f"Added {name} with {record_action} Grade {grade_level} record. Logged {changes_logged + 1} change(s).")
    return redirect(url_for("student_history", lrn=lrn))


@app.route("/student/<lrn>/record/add", methods=["POST"])
@admin_required
def add_student_year_record(lrn):
    ensure_schema()

    if not LRN_PATTERN.match(lrn):
        flash("Invalid LRN. Use exactly 12 digits.")
        return redirect(url_for("students"))

    gender = normalize_gender(request.form.get("gender", ""))
    school_year = request.form.get("school_year", "").strip()
    grade_level = request.form.get("grade_level", "").strip()
    status = request.form.get("status", "").strip()
    remarks = normalize_remarks(request.form.get("remarks", ""))
    confirmation = request.form.get("confirmation", "").strip()

    if confirmation != "ADD RECORD":
        flash("Type ADD RECORD to confirm adding a new school year record.")
        return redirect(url_for("student_history", lrn=lrn))

    if gender not in {"MALE", "FEMALE"}:
        flash("Select a valid sex value.")
        return redirect(url_for("student_history", lrn=lrn))

    if not is_valid_school_year(school_year):
        flash("Use school year format YYYY-YYYY with consecutive years before adding a year record.")
        return redirect(url_for("student_history", lrn=lrn))

    if not grade_level.isdigit() or int(grade_level) not in SUPPORTED_GRADES:
        flash("Select a valid Grade 7 to Grade 10 record.")
        return redirect(url_for("student_history", lrn=lrn))

    if status not in EDITABLE_RECORD_STATUSES:
        flash("Select a valid student status.")
        return redirect(url_for("student_history", lrn=lrn))

    grade_level = int(grade_level)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM students WHERE lrn = %s", (lrn,))
    student = cursor.fetchone()

    if student is None:
        cursor.close()
        conn.close()
        flash("No student found for that LRN.")
        return redirect(url_for("students"))

    cursor.execute(
        """
        SELECT id, grade_level
        FROM student_records
        WHERE lrn = %s
        AND school_year = %s
        LIMIT 1
        """,
        (lrn, school_year),
    )
    existing = cursor.fetchone()

    if existing:
        cursor.close()
        conn.close()
        flash(f"Duplicate school year rejected. {lrn} already has a Grade {existing[1]} record in {school_year}.")
        return redirect(url_for("student_history", lrn=lrn))

    changed_by = session.get("username")
    cursor.execute(
        """
        INSERT INTO student_records (lrn, school_year, grade_level, gender, status, remarks)
        VALUES (%s, %s, %s, %s, %s, %s)
        """,
        (lrn, school_year, grade_level, gender, status, remarks),
    )
    log_change(cursor, lrn, "manual.student_record", "", f"Added Grade {grade_level} record for {school_year}", school_year, grade_level, changed_by)
    conn.commit()
    cursor.close()
    conn.close()

    flash(f"Added Grade {grade_level} record for {student[0]} in {school_year}.")
    return redirect(url_for("student_history", lrn=lrn))


@app.route("/reports")
@login_required
def reports():
    start_year = request.args.get("start_year", "").strip()
    start_grade = request.args.get("start_grade", "7").strip()
    if start_year:
        return redirect(url_for("cohort_tracking", start_year=start_year, start_grade=start_grade))
    return redirect(url_for("cohort_tracking"))


@app.route("/reports/export")
@login_required
def export_report():
    ensure_schema()

    start_year = request.args.get("start_year", "").strip()
    start_grade = request.args.get("start_grade", "7").strip()
    if not start_year:
        flash("Select a cohort starting school year before exporting.")
        return redirect(url_for("cohort_tracking"))
    if not is_valid_school_year(start_year):
        flash("Use a valid starting school year in YYYY-YYYY format before exporting.")
        return redirect(url_for("cohort_tracking"))
    if not start_grade.isdigit() or int(start_grade) not in SUPPORTED_GRADES:
        flash("Select a valid Grade 7 to Grade 10 entry level before exporting.")
        return redirect(url_for("cohort_tracking", start_year=start_year))

    start_grade = int(start_grade)

    conn = get_db_connection()
    cursor = conn.cursor()
    report = build_grade7_cohort_report(cursor, start_year, start_grade)
    cursor.close()
    conn.close()

    workbook = build_grade7_cohort_report_workbook(report)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"grade-{start_grade}-entry-cohort-progression-report-{start_year}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/reports/print")
@login_required
def print_report():
    ensure_schema()

    start_year = request.args.get("start_year", "").strip()
    start_grade = request.args.get("start_grade", "7").strip()
    if not start_year:
        flash("Select a cohort starting school year before printing.")
        return redirect(url_for("cohort_tracking"))
    if not is_valid_school_year(start_year):
        flash("Use a valid starting school year in YYYY-YYYY format before printing.")
        return redirect(url_for("cohort_tracking"))
    if not start_grade.isdigit() or int(start_grade) not in SUPPORTED_GRADES:
        flash("Select a valid Grade 7 to Grade 10 entry level before printing.")
        return redirect(url_for("cohort_tracking", start_year=start_year))

    start_grade = int(start_grade)

    conn = get_db_connection()
    cursor = conn.cursor()
    report = build_grade7_cohort_report(cursor, start_year, start_grade)
    cursor.close()
    conn.close()

    return render_template(
        "print_report.html",
        report=report,
    )


@app.route("/templates/<path:filename>")
def template_assets(filename):
    return send_from_directory("templates", filename)


@app.route("/rsc/<path:filename>")
def resources(filename):
    return send_from_directory("rsc", filename)


@app.route("/student/search")
@login_required
def search_student():
    lrn = request.args.get("lrn", "").strip()

    if not LRN_PATTERN.match(lrn):
        flash("Enter a valid 12-digit LRN to view student history.")
        return redirect(url_for("students"))

    return redirect(url_for("student_history", lrn=lrn))


@app.route("/student/<lrn>")
@login_required
def student_history(lrn):
    ensure_schema()

    if not LRN_PATTERN.match(lrn):
        flash("Invalid LRN. Use exactly 12 digits.")
        return redirect(url_for("students"))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT
            s.lrn,
            s.name,
            COALESCE(
                s.gender,
                (
                    SELECT r.gender
                    FROM student_records r
                    WHERE r.lrn = s.lrn
                    ORDER BY r.school_year DESC, r.grade_level DESC
                    LIMIT 1
                )
            ),
            s.updated_at
        FROM students s
        WHERE s.lrn = %s
        """,
        (lrn,),
    )
    student = cursor.fetchone()

    if student is None:
        cursor.close()
        conn.close()
        flash("No student found for that LRN.")
        return redirect(url_for("students"))

    cursor.execute(
        """
        SELECT school_year, grade_level, gender, status, remarks, updated_at
        FROM student_records
        WHERE lrn = %s
        ORDER BY school_year, grade_level
        """,
        (lrn,),
    )
    history = [
        {
            "school_year": school_year,
            "grade_level": grade_level,
            "gender": gender,
            "status": status,
            "status_label": humanize_status(status),
            "remarks": format_remarks(status, remarks),
            "raw_remarks": normalize_remarks(remarks),
            "updated_at": updated_at,
        }
        for school_year, grade_level, gender, status, remarks, updated_at in cursor.fetchall()
    ]

    cursor.execute(
        """
        SELECT field_name, old_value, new_value, school_year, grade_level, changed_by, changed_at
        FROM student_change_logs
        WHERE lrn = %s
        ORDER BY changed_at DESC
        """,
        (lrn,),
    )
    changes = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        "student_history.html",
        student=student,
        history=history,
        changes=changes,
        grades=SUPPORTED_GRADES,
        editable_statuses=EDITABLE_RECORD_STATUSES
    )


@app.route("/student/<lrn>/record/update", methods=["POST"])
@admin_required
def update_student_record(lrn):
    ensure_schema()

    if not LRN_PATTERN.match(lrn):
        flash("Invalid LRN. Use exactly 12 digits.")
        return redirect(url_for("students"))

    school_year = request.form.get("school_year", "").strip()
    grade_level = request.form.get("grade_level", "").strip()
    status = request.form.get("status", "").strip()
    remarks = normalize_remarks(request.form.get("remarks", ""))

    if not is_valid_school_year(school_year):
        flash("Use school year format YYYY-YYYY with consecutive years before updating a record.")
        return redirect(url_for("student_history", lrn=lrn))

    if not grade_level.isdigit() or int(grade_level) not in SUPPORTED_GRADES:
        flash("Select a valid Grade 7 to Grade 10 record.")
        return redirect(url_for("student_history", lrn=lrn))

    if status not in EDITABLE_RECORD_STATUSES:
        flash("Select a valid student status.")
        return redirect(url_for("student_history", lrn=lrn))

    grade_level = int(grade_level)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT id, status, remarks
        FROM student_records
        WHERE lrn = %s
        AND school_year = %s
        AND grade_level = %s
        LIMIT 1
        """,
        (lrn, school_year, grade_level),
    )
    existing = cursor.fetchone()

    if existing is None:
        cursor.close()
        conn.close()
        flash("No matching progression record was found.")
        return redirect(url_for("student_history", lrn=lrn))

    record_id, old_status, old_remarks = existing
    changes_logged = 0
    changed_by = session.get("username")

    if log_change(cursor, lrn, "record.status", old_status, status, school_year, grade_level, changed_by=changed_by):
        changes_logged += 1
    if log_change(cursor, lrn, "record.remarks", old_remarks, remarks, school_year, grade_level, changed_by=changed_by):
        changes_logged += 1

    if changes_logged:
        cursor.execute(
            """
            UPDATE student_records
            SET status = %s, remarks = %s
            WHERE id = %s
            """,
            (status, remarks, record_id),
        )
        conn.commit()
        flash("Student record updated and logged.")
    else:
        flash("No changes were made.")

    cursor.close()
    conn.close()
    return redirect(url_for("student_history", lrn=lrn))


@app.route("/student/<lrn>/delete", methods=["POST"])
@admin_required
def delete_student(lrn):
    ensure_schema()

    if not LRN_PATTERN.match(lrn):
        flash("Invalid LRN. Use exactly 12 digits.")
        return redirect(url_for("students"))

    confirmation = request.form.get("confirmation", "").strip()
    if confirmation != "DELETE STUDENT":
        flash("Type DELETE STUDENT to confirm student deletion.")
        return redirect(url_for("student_history", lrn=lrn))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM students WHERE lrn = %s", (lrn,))
    student = cursor.fetchone()

    if student is None:
        cursor.close()
        conn.close()
        flash("No student found for that LRN.")
        return redirect(url_for("students"))

    cursor.execute("SELECT COUNT(*) FROM student_records WHERE lrn = %s", (lrn,))
    record_count = cursor.fetchone()[0] or 0

    cursor.execute("DELETE FROM student_change_logs WHERE lrn = %s", (lrn,))
    cursor.execute("DELETE FROM student_records WHERE lrn = %s", (lrn,))
    cursor.execute("DELETE FROM students WHERE lrn = %s", (lrn,))
    conn.commit()
    cursor.close()
    conn.close()

    flash(f"Deleted {student[0]} and {record_count} progression record(s).")
    return redirect(url_for("students"))


@app.route("/cohort-tracking")
@login_required
def cohort_tracking():
    ensure_schema()

    start_year = request.args.get("start_year", "").strip()
    start_grade = request.args.get("start_grade", "7").strip()
    cohort_rows = []
    summary = {
        "total": 0,
        "straight_path": 0,
        "completed": 0,
        "delayed": 0,
        "repeated": 0,
        "transfer_in": 0,
        "transfer_out": 0,
        "dropped": 0,
        "incomplete": 0,
        "for_review": 0,
    }
    expected_path = []

    if start_year:
        if not is_valid_school_year(start_year):
            flash("Use school year format YYYY-YYYY with consecutive years for cohort tracking.")
            return redirect(url_for("cohort_tracking"))

        if not start_grade.isdigit() or int(start_grade) not in SUPPORTED_GRADES:
            flash("Select a valid starting grade from Grade 7 to Grade 10.")
            return redirect(url_for("cohort_tracking"))

        start_grade = int(start_grade)
        expected_path = build_expected_path(start_year, start_grade)

        conn = get_db_connection()
        cursor = conn.cursor()
        cohort_rows, summary = build_cohort_tracking(cursor, start_year, start_grade, expected_path)
        cursor.close()
        conn.close()

    return render_template(
        "co_tracking.html",
        grades=SUPPORTED_GRADES,
        start_year=start_year,
        start_grade=int(start_grade) if str(start_grade).isdigit() else 7,
        expected_path=expected_path,
        cohort_rows=cohort_rows,
        summary=summary,
    )


def get_metrics():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM students")
    total = cursor.fetchone()[0]

    cursor.close()
    conn.close()

    return total


def get_dashboard_stats(cursor):
    cursor.execute("SELECT COUNT(*) FROM students")
    total_students = cursor.fetchone()[0] or 0

    cursor.execute("SELECT COUNT(*) FROM student_records")
    total_records = cursor.fetchone()[0] or 0

    cursor.execute(
        """
        SELECT
            SUM(CASE WHEN status = 'ENROLLED' THEN 1 ELSE 0 END),
            SUM(CASE WHEN status = 'TRANSFER_IN' THEN 1 ELSE 0 END),
            SUM(CASE WHEN status = 'PENDING_TRANSFER_IN' THEN 1 ELSE 0 END),
            SUM(CASE WHEN status = 'TRANSFER_OUT' THEN 1 ELSE 0 END)
        FROM student_records
        """
    )
    enrolled, transfer_in, pending_transfer_in, transfer_out = cursor.fetchone()

    cursor.execute(
        """
        SELECT COUNT(*)
        FROM (
            SELECT lrn, grade_level
            FROM student_records
            GROUP BY lrn, grade_level
            HAVING COUNT(DISTINCT school_year) > 1
        ) repeated
        """
    )
    repeaters = cursor.fetchone()[0] or 0

    cursor.execute("SELECT COUNT(DISTINCT lrn) FROM student_records WHERE grade_level = 10")
    grade_10 = cursor.fetchone()[0] or 0

    return {
        "total_students": total_students,
        "total_records": total_records,
        "active_students": enrolled or 0,
        "transfer_in": transfer_in or 0,
        "pending_transfer_in": pending_transfer_in or 0,
        "transfer_out": transfer_out or 0,
        "repeaters": repeaters,
        "completed": grade_10,
    }


def get_grade_distribution(cursor):
    cursor.execute(
        """
        SELECT grade_level, COUNT(DISTINCT lrn)
        FROM student_records
        WHERE grade_level IN (7, 8, 9, 10)
        GROUP BY grade_level
        """
    )
    counts = {grade: count for grade, count in cursor.fetchall()}

    return [
        {
            "grade": grade,
            "count": counts.get(grade, 0),
            "label": "Tracked" if counts.get(grade, 0) else "No records",
        }
        for grade in SUPPORTED_GRADES
    ]


def get_recent_changes(cursor, limit=5):
    limit = max(1, min(int(limit), 20))
    cursor.execute(
        f"""
        SELECT lrn, field_name, old_value, new_value, school_year, grade_level, changed_at
        FROM student_change_logs
        ORDER BY changed_at DESC
        LIMIT {limit}
        """
    )

    return [
        {
            "lrn": lrn,
            "field_name": field_name,
            "old_value": old_value,
            "new_value": new_value,
            "school_year": school_year,
            "grade_level": grade_level,
            "changed_at": changed_at,
        }
        for lrn, field_name, old_value, new_value, school_year, grade_level, changed_at in cursor.fetchall()
    ]


def build_student_record_filters(filters):
    where = " WHERE 1=1"
    params = []

    if filters.get("q"):
        where += " AND (s.lrn LIKE %s OR s.name LIKE %s)"
        search = f"%{filters['q']}%"
        params.extend([search, search])

    if filters.get("grade") and filters["grade"].isdigit() and int(filters["grade"]) in SUPPORTED_GRADES:
        where += " AND r.grade_level = %s"
        params.append(int(filters["grade"]))

    if filters.get("status"):
        where += " AND r.status = %s"
        params.append(filters["status"])

    if filters.get("year") and is_valid_school_year(filters["year"]):
        where += " AND r.school_year = %s"
        params.append(filters["year"])

    return where, params


def count_student_records(cursor, filters):
    where, params = build_student_record_filters(filters)
    query = """
        SELECT COUNT(*)
        FROM students s
        JOIN student_records r ON s.lrn = r.lrn
    """ + where
    cursor.execute(query, params)
    return cursor.fetchone()[0] or 0


def get_student_records(cursor, filters, page=1, per_page=PER_PAGE):
    where, params = build_student_record_filters(filters)
    offset = (page - 1) * per_page
    query = """
        SELECT s.lrn, s.name, r.gender, r.grade_level, r.school_year, r.status, r.remarks
        FROM students s
        JOIN student_records r ON s.lrn = r.lrn
    """ + where + " ORDER BY r.school_year DESC, r.grade_level, s.name LIMIT %s OFFSET %s"
    cursor.execute(query, params + [per_page, offset])

    return [
        {
            "lrn": lrn,
            "name": name,
            "gender": gender,
            "grade_level": grade_level,
            "school_year": school_year,
            "status": status,
            "status_label": humanize_status(status),
            "remarks": format_remarks(status, remarks),
        }
        for lrn, name, gender, grade_level, school_year, status, remarks in cursor.fetchall()
    ]


def get_school_years(cursor):
    cursor.execute(
        """
        SELECT DISTINCT school_year
        FROM student_records
        ORDER BY school_year DESC
        """
    )
    return [row[0] for row in cursor.fetchall()]


def get_report_records(cursor, school_year=""):
    query = """
        SELECT s.lrn, s.name, r.gender, r.grade_level, r.school_year, r.status, r.remarks
        FROM students s
        JOIN student_records r ON s.lrn = r.lrn
        WHERE 1=1
    """
    params = []

    if school_year:
        query += " AND r.school_year = %s"
        params.append(school_year)

    query += " ORDER BY r.school_year DESC, r.grade_level, s.name"
    cursor.execute(query, params)

    return [
        {
            "lrn": lrn,
            "name": name,
            "gender": gender or "",
            "grade_level": grade_level,
            "school_year": record_year,
            "status": status,
            "status_label": humanize_status(status),
            "remarks": format_remarks(status, remarks),
        }
        for lrn, name, gender, grade_level, record_year, status, remarks in cursor.fetchall()
    ]


def humanize_status(status):
    labels = {
        "ENROLLED": "Enrolled",
        "TRANSFER_IN": "Transferred In",
        "PENDING_TRANSFER_IN": "Pending Transfer-In",
        "TRANSFER_OUT": "Transferred Out",
        "MISSING": "Missing",
        "REPEATED": "Repeated",
        "COMPLETED": "Completed",
        "DELAYED_COMPLETED": "Completed - Delayed",
        "STRAIGHT_PATH": "Regular",
        "INCOMPLETE": "Incomplete",
    }
    return labels.get(status, str(status or "").replace("_", " ").title())


def format_remarks(status, remarks):
    clean_remarks = normalize_remarks(remarks)
    status_label = humanize_status(status)

    abbreviated_remarks = {
        "T/I": "Transferred In",
        "T-I": "Transferred In",
        "TI": "Transferred In",
        "TRANSFER IN": "Transferred In",
        "PENDING T/I": "Pending Transfer-In",
        "PENDING TI": "Pending Transfer-In",
        "T/O": "Transferred Out",
        "T-O": "Transferred Out",
        "TO": "Transferred Out",
        "TRANSFER OUT": "Transferred Out",
    }

    if clean_remarks.upper() in abbreviated_remarks:
        return abbreviated_remarks[clean_remarks.upper()]

    if status in {"TRANSFER_IN", "PENDING_TRANSFER_IN", "TRANSFER_OUT"} and not clean_remarks:
        return status_label

    return clean_remarks


def build_student_timelines(cursor):
    cursor.execute(
        """
        SELECT s.lrn, s.name, r.school_year, r.grade_level, r.status, r.remarks
        FROM students s
        JOIN student_records r ON s.lrn = r.lrn
        WHERE r.grade_level IN (7, 8, 9, 10)
        ORDER BY s.lrn, r.school_year, r.grade_level
        """
    )

    timelines = {}
    for lrn, name, school_year, grade_level, status, remarks in cursor.fetchall():
        timelines.setdefault(lrn, {"lrn": lrn, "name": name, "records": []})
        timelines[lrn]["records"].append(
            {
                "school_year": school_year,
                "grade_level": grade_level,
                "status": status,
                "status_label": humanize_status(status),
                "remarks": format_remarks(status, remarks),
            }
        )

    return timelines


def get_at_risk_students(cursor):
    timelines = build_student_timelines(cursor)
    existing_years = set(get_school_years(cursor))
    at_risk = []

    for timeline in timelines.values():
        records = timeline["records"]
        latest = records[-1]

        if latest["status"] == "TRANSFER_OUT":
            at_risk.append(
                {
                    "lrn": timeline["lrn"],
                    "name": timeline["name"],
                    "last_grade": latest["grade_level"],
                    "last_year": latest["school_year"],
                    "reason": f"Transferred out after Grade {latest['grade_level']}",
                    "remarks": latest["remarks"] or "Transferred Out",
                }
            )
            continue

        if latest["grade_level"] < 10 and next_school_year(latest["school_year"]) in existing_years:
            if latest["status"] == "PENDING_TRANSFER_IN":
                reason = f"Pending Transfer-In / Needs Verification after Grade {latest['grade_level']}"
                remarks = latest["remarks"] or "Pending Transfer-In"
            else:
                reason = f"Did not appear after Grade {latest['grade_level']}"
                remarks = latest["remarks"] or "-"

            at_risk.append(
                {
                    "lrn": timeline["lrn"],
                    "name": timeline["name"],
                    "last_grade": latest["grade_level"],
                    "last_year": latest["school_year"],
                    "reason": reason,
                    "remarks": remarks,
                }
            )

    return at_risk


def get_csr_breakdown(at_risk_list):
    breakdown = {
        grade: {
            "grade": grade,
            "transferred_out": 0,
            "pending_verification": 0,
            "missing_after": 0,
            "total_left": 0,
        }
        for grade in SUPPORTED_GRADES
    }

    for student in at_risk_list:
        grade = student["last_grade"]
        if grade not in breakdown:
            continue

        if "Transferred out" in student["reason"]:
            breakdown[grade]["transferred_out"] += 1
        elif "Pending Transfer-In" in student["reason"]:
            breakdown[grade]["pending_verification"] += 1
        else:
            breakdown[grade]["missing_after"] += 1

        breakdown[grade]["total_left"] += 1

    return list(breakdown.values())


def style_report_sheet(sheet):
    dark_blue = "12376F"
    light_blue = "EAF1FF"
    border_color = "D9E2F3"
    thin = Side(style="thin", color=border_color)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=dark_blue)
        cell.font = Font(color="FFFFFF", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")

    for row_number in (5, 12):
        for cell in sheet[row_number]:
            cell.fill = PatternFill("solid", fgColor=light_blue)
            cell.font = Font(bold=True, color=dark_blue)

    for column_index, column_cells in enumerate(sheet.columns, start=1):
        column = get_column_letter(column_index)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column].width = min(max(max_length + 3, 12), 42)

    sheet.freeze_panes = "A13"


def build_grade7_cohort_report_workbook(report):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cohort Report"

    sheet.merge_cells("A1:G1")
    sheet["A1"] = report["title"]
    sheet["A2"] = "Entry Cohort Start"
    sheet["B2"] = report["start_year"]
    sheet["A3"] = "Expected Grade 10 Year"
    sheet["B3"] = report["end_year"]

    summary_rows = [
        ("Summary", "Value", "", "Indicator", "Value", "", "Meaning"),
        (f"Original Grade {report['start_grade']} Entry Cohort", report["summary"]["total"], "", "On-Time Completion Rate", f"{report['rates']['on_time_completion']}%", "", "Completed Grade 10 on the expected year"),
        ("On-Time Completed", report["summary"]["on_time"], "", "Overall Completion Rate", f"{report['rates']['overall_completion']}%", "", "Completed Grade 10 even if delayed"),
        ("Completed Grade 10", report["summary"]["completed"], "", "Delayed / Repeated", report["summary"]["delayed"], "", "Learners with repetition or late completion"),
        ("Transfer-Out", report["summary"]["transfer_out"], "", "For Review", report["summary"]["for_review"], "", "Records needing verification"),
        ("Incomplete", report["summary"]["incomplete"], "", "", "", "", ""),
    ]
    for offset, row in enumerate(summary_rows, start=5):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=offset, column=column, value=value)

    path_row = 12
    path_headers = ["Expected Path"] + [f"Grade {step['grade']}" for step in report["expected_path"]]
    for column, value in enumerate(path_headers, start=1):
        sheet.cell(row=path_row, column=column, value=value)
    sheet.cell(row=path_row + 1, column=1, value="School Year")
    for column, step in enumerate(report["expected_path"], start=2):
        sheet.cell(row=path_row + 1, column=column, value=step["school_year"])

    table_start = 16
    path_columns = [f"Grade {step['grade']} ({step['school_year']})" for step in report["expected_path"]]
    headers = ["LRN", "Student Name"] + path_columns + ["Final Result", "Review Note"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=table_start, column=column, value=header)

    for row_index, row_data in enumerate(report["rows"], start=table_start + 1):
        sheet.cell(row=row_index, column=1, value=row_data["lrn"])
        sheet.cell(row=row_index, column=2, value=row_data["name"])
        for column, cell in enumerate(row_data["path"], start=3):
            value = cell["status_label"]
            if cell["remarks"]:
                value = f"{value} - {cell['remarks']}"
            sheet.cell(row=row_index, column=column, value=value)
        sheet.cell(row=row_index, column=3 + len(report["expected_path"]), value=row_data["result_label"])
        sheet.cell(row=row_index, column=4 + len(report["expected_path"]), value=row_data["reason"] or "")

    if not report["rows"]:
        sheet.cell(row=table_start + 1, column=1, value=f"No Grade {report['start_grade']} learners found for this starting school year.")

    style_report_sheet(sheet)

    review_sheet = workbook.create_sheet("For Review")
    review_sheet.merge_cells("A1:F1")
    review_sheet["A1"] = "Learners for Review"
    review_headers = ["LRN", "Student Name", "Last Grade Seen", "Last School Year", "Reason", "Remarks"]
    for column, header in enumerate(review_headers, start=1):
        review_sheet.cell(row=5, column=column, value=header)

    for row_index, learner in enumerate(report["review_learners"], start=6):
        review_sheet.cell(row=row_index, column=1, value=learner["lrn"])
        review_sheet.cell(row=row_index, column=2, value=learner["name"])
        review_sheet.cell(row=row_index, column=3, value=f"Grade {learner['last_grade']}")
        review_sheet.cell(row=row_index, column=4, value=learner["last_year"])
        review_sheet.cell(row=row_index, column=5, value=learner["reason"])
        review_sheet.cell(row=row_index, column=6, value=learner["remarks"])

    if not report["review_learners"]:
        review_sheet.cell(row=6, column=1, value="No learners need review for this cohort.")

    style_report_sheet(review_sheet)

    breakdown_sheet = workbook.create_sheet("Review Breakdown")
    breakdown_sheet.merge_cells("A1:E1")
    breakdown_sheet["A1"] = "Progression Review Breakdown"
    breakdown_headers = ["Grade Level", "Transferred Out", "Delayed / Repeated", "Missing / Incomplete", "Total for Review"]
    for column, header in enumerate(breakdown_headers, start=1):
        breakdown_sheet.cell(row=5, column=column, value=header)

    for row_index, item in enumerate(report["breakdown"], start=6):
        breakdown_sheet.cell(row=row_index, column=1, value=f"Grade {item['grade']}")
        breakdown_sheet.cell(row=row_index, column=2, value=item["transferred_out"])
        breakdown_sheet.cell(row=row_index, column=3, value=item["repeated_or_delayed"])
        breakdown_sheet.cell(row=row_index, column=4, value=item["missing"])
        breakdown_sheet.cell(row=row_index, column=5, value=item["for_review"])

    style_report_sheet(breakdown_sheet)
    return workbook


def compute_promotion(year1, year2, grade_from, grade_to):
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT COUNT(*)
        FROM student_records r1
        JOIN student_records r2
        ON r1.lrn = r2.lrn
        WHERE r1.school_year = %s
        AND r2.school_year = %s
        AND r1.grade_level = %s
        AND r2.grade_level = %s
        """,
        (year1, year2, grade_from, grade_to),
    )
    promoted = cursor.fetchone()[0]

    cursor.execute(
        """
        SELECT COUNT(DISTINCT lrn)
        FROM student_records
        WHERE school_year = %s
        AND grade_level = %s
        """,
        (year1, grade_from),
    )
    total = cursor.fetchone()[0]

    cursor.execute(
        """
        SELECT COUNT(*)
        FROM student_records r1
        JOIN student_records r2
        ON r1.lrn = r2.lrn
        WHERE r1.school_year = %s
        AND r2.school_year = %s
        AND r1.grade_level = %s
        AND r2.grade_level = %s
        """,
        (year1, year2, grade_from, grade_from),
    )
    repeated = cursor.fetchone()[0]

    dropped = total - (promoted + repeated)
    rate = (promoted / total * 100) if total > 0 else 0

    cursor.close()
    conn.close()

    return {
        "promoted": promoted,
        "repeated": repeated,
        "dropped": dropped,
        "rate": round(rate, 2),
    }


def next_school_year(school_year):
    start, end = school_year.split("-")
    return f"{int(start) + 1}-{int(end) + 1}"


def build_expected_path(start_year, start_grade):
    path = []
    current_year = start_year

    for grade in range(start_grade, 11):
        path.append({"grade": grade, "school_year": current_year})
        current_year = next_school_year(current_year)

    return path


def summarize_cohort_status(path_cells, records=None):
    records = records or []
    statuses = [cell["status"] for cell in path_cells if cell["status"] != "MISSING"]
    all_statuses = statuses + [record[2] for record in records]
    grades = [cell["actual_grade"] for cell in path_cells if cell["actual_grade"] is not None]
    all_grades = [record[1] for record in records]
    has_missing = any(cell["status"] == "MISSING" for cell in path_cells)
    has_grade_10 = 10 in all_grades or (
        path_cells and path_cells[-1]["status"] != "MISSING" and path_cells[-1]["expected_grade"] == 10
    )

    if "TRANSFER_OUT" in all_statuses:
        return "TRANSFER_OUT"

    if has_grade_10:
        return "DELAYED_COMPLETED" if has_missing else "COMPLETED"

    if "TRANSFER_IN" in all_statuses or "PENDING_TRANSFER_IN" in all_statuses:
        return "TRANSFER_IN"

    if len(grades) != len(set(grades)) or len(all_grades) != len(set(all_grades)):
        return "REPEATED"

    if has_missing:
        return "INCOMPLETE"

    return "STRAIGHT_PATH"


def build_cohort_tracking(cursor, start_year, start_grade, expected_path):
    cursor.execute(
        """
        SELECT DISTINCT s.lrn, s.name
        FROM students s
        JOIN student_records r ON s.lrn = r.lrn
        WHERE r.school_year = %s
        AND r.grade_level = %s
        ORDER BY s.name
        """,
        (start_year, start_grade),
    )
    cohort_students = cursor.fetchall()

    rows = []
    summary = {
        "total": len(cohort_students),
        "straight_path": 0,
        "completed": 0,
        "delayed": 0,
        "repeated": 0,
        "transfer_in": 0,
        "transfer_out": 0,
        "dropped": 0,
        "incomplete": 0,
        "for_review": 0,
    }

    for lrn, name in cohort_students:
        cursor.execute(
            """
            SELECT school_year, grade_level, status, remarks
            FROM student_records
            WHERE lrn = %s
            AND grade_level BETWEEN %s AND 10
            ORDER BY school_year, grade_level
            """,
            (lrn, start_grade),
        )
        records = cursor.fetchall()
        records_by_year_grade = {
            (school_year, grade_level): {
                "status": status,
                "remarks": remarks,
                "grade": grade_level,
            }
            for school_year, grade_level, status, remarks in records
        }

        path_cells = []

        for step in expected_path:
            record = records_by_year_grade.get((step["school_year"], step["grade"]))

            if record:
                path_cells.append(
                    {
                        "school_year": step["school_year"],
                        "expected_grade": step["grade"],
                        "actual_grade": record["grade"],
                        "status": record["status"],
                        "status_label": humanize_status(record["status"]),
                        "remarks": format_remarks(record["status"], record["remarks"]),
                    }
                )
            else:
                path_cells.append(
                    {
                        "school_year": step["school_year"],
                        "expected_grade": step["grade"],
                        "actual_grade": None,
                        "status": "MISSING",
                        "status_label": humanize_status("MISSING"),
                        "remarks": "",
                    }
                )

        result = summarize_cohort_status(path_cells, records)

        if result == "COMPLETED":
            summary["completed"] += 1
            summary["straight_path"] += 1
        elif result == "DELAYED_COMPLETED":
            summary["completed"] += 1
            summary["delayed"] += 1
        elif result == "STRAIGHT_PATH":
            summary["straight_path"] += 1
        elif result == "REPEATED":
            summary["repeated"] += 1
        elif result == "TRANSFER_IN":
            summary["transfer_in"] += 1
        elif result == "TRANSFER_OUT":
            summary["transfer_out"] += 1
        else:
            summary["incomplete"] += 1

        if result == "INCOMPLETE" and any(cell["status"] == "MISSING" for cell in path_cells):
            summary["dropped"] += 1

        if result not in {"COMPLETED", "STRAIGHT_PATH"}:
            summary["for_review"] += 1

        rows.append(
            {
                "lrn": lrn,
                "name": name,
                "path": path_cells,
                "result": result,
                "result_label": humanize_status(result),
            }
        )

    return rows, summary


def build_grade7_cohort_report(cursor, start_year, start_grade=7):
    expected_path = build_expected_path(start_year, start_grade)
    end_year = expected_path[-1]["school_year"]

    cursor.execute(
        """
        SELECT DISTINCT s.lrn, s.name
        FROM students s
        JOIN student_records r ON s.lrn = r.lrn
        WHERE r.school_year = %s
        AND r.grade_level = %s
        ORDER BY s.name
        """,
        (start_year, start_grade),
    )
    cohort_students = cursor.fetchall()

    summary = {
        "total": len(cohort_students),
        "on_time": 0,
        "completed": 0,
        "delayed": 0,
        "repeated": 0,
        "transfer_out": 0,
        "incomplete": 0,
        "for_review": 0,
    }
    breakdown = {
        grade: {
            "grade": grade,
            "missing": 0,
            "repeated_or_delayed": 0,
            "transferred_out": 0,
            "for_review": 0,
        }
        for grade in SUPPORTED_GRADES
    }
    review_learners = []
    rows = []

    for lrn, name in cohort_students:
        cursor.execute(
            """
            SELECT school_year, grade_level, status, remarks
            FROM student_records
            WHERE lrn = %s
            AND grade_level BETWEEN %s AND 10
            ORDER BY school_year, grade_level
            """,
            (lrn, start_grade),
        )
        records = [
            {
                "school_year": school_year,
                "grade_level": grade_level,
                "status": status,
                "remarks": remarks,
            }
            for school_year, grade_level, status, remarks in cursor.fetchall()
        ]
        records_by_year_grade = {
            (record["school_year"], record["grade_level"]): record
            for record in records
        }
        grade_years = {}
        for record in records:
            grade_years.setdefault(record["grade_level"], set()).add(record["school_year"])

        path_cells = []
        missing_steps = []
        for step in expected_path:
            record = records_by_year_grade.get((step["school_year"], step["grade"]))
            if record:
                path_cells.append(
                    {
                        "school_year": step["school_year"],
                        "expected_grade": step["grade"],
                        "status": record["status"],
                        "status_label": humanize_status(record["status"]),
                        "remarks": format_remarks(record["status"], record["remarks"]),
                    }
                )
            else:
                missing_steps.append(step)
                path_cells.append(
                    {
                        "school_year": step["school_year"],
                        "expected_grade": step["grade"],
                        "status": "MISSING",
                        "status_label": humanize_status("MISSING"),
                        "remarks": "",
                    }
                )

        has_grade10 = any(record["grade_level"] == 10 for record in records)
        has_expected_grade10 = (end_year, 10) in records_by_year_grade
        has_transfer_out = any(record["status"] == "TRANSFER_OUT" for record in records)
        has_pending_transfer = any(record["status"] == "PENDING_TRANSFER_IN" for record in records)
        has_repetition = any(len(years) > 1 for years in grade_years.values())
        on_time = has_expected_grade10 and not missing_steps and not has_transfer_out
        delayed = has_grade10 and not on_time

        if has_transfer_out:
            result = "TRANSFER_OUT"
            reason = "Transferred out before completing the expected path"
        elif on_time:
            result = "COMPLETED"
            reason = ""
        elif delayed:
            result = "DELAYED_COMPLETED"
            reason = "Reached Grade 10 later than the expected path"
        elif has_repetition:
            result = "REPEATED"
            reason = "Repeated or delayed in one grade level"
        else:
            result = "INCOMPLETE"
            reason = "Missing expected progression record"

        if has_pending_transfer and result not in {"COMPLETED", "DELAYED_COMPLETED"}:
            reason = "Pending transfer-in record needs verification"

        if on_time:
            summary["on_time"] += 1
        if has_grade10:
            summary["completed"] += 1
        if delayed or has_repetition:
            summary["delayed"] += 1
        if has_repetition:
            summary["repeated"] += 1
        if has_transfer_out:
            summary["transfer_out"] += 1
        if result == "INCOMPLETE":
            summary["incomplete"] += 1
        if result in {"TRANSFER_OUT", "REPEATED", "INCOMPLETE", "DELAYED_COMPLETED"} or has_pending_transfer:
            summary["for_review"] += 1

        if has_transfer_out:
            transfer_record = next((record for record in records if record["status"] == "TRANSFER_OUT"), records[-1])
            grade = transfer_record["grade_level"]
            breakdown[grade]["transferred_out"] += 1
            breakdown[grade]["for_review"] += 1
        if has_repetition or delayed:
            latest_grade = max(record["grade_level"] for record in records)
            breakdown[latest_grade]["repeated_or_delayed"] += 1
            breakdown[latest_grade]["for_review"] += 1
        if missing_steps and not has_grade10 and not has_transfer_out:
            grade = missing_steps[0]["grade"]
            breakdown[grade]["missing"] += 1
            breakdown[grade]["for_review"] += 1

        if reason:
            latest = records[-1] if records else {"grade_level": start_grade, "school_year": start_year, "remarks": ""}
            review_learners.append(
                {
                    "lrn": lrn,
                    "name": name,
                    "last_grade": latest["grade_level"],
                    "last_year": latest["school_year"],
                    "reason": reason,
                    "remarks": format_remarks(latest.get("status"), latest.get("remarks")) or "-",
                }
            )

        rows.append(
            {
                "lrn": lrn,
                "name": name,
                "path": path_cells,
                "result": result,
                "result_label": humanize_status(result),
                "reason": reason,
            }
        )

    rates = {
        "on_time_completion": round((summary["on_time"] / summary["total"]) * 100, 2) if summary["total"] else 0,
        "overall_completion": round((summary["completed"] / summary["total"]) * 100, 2) if summary["total"] else 0,
    }

    return {
        "start_year": start_year,
        "end_year": end_year,
        "start_grade": start_grade,
        "title": f"Grade {start_grade} Entry Cohort Progression Report: {start_year} to {end_year}",
        "expected_path": expected_path,
        "summary": summary,
        "rates": rates,
        "breakdown": list(breakdown.values()),
        "review_learners": review_learners,
        "rows": rows,
    }


def ensure_schema():
    global SCHEMA_READY

    if SCHEMA_READY:
        return

    conn = get_db_connection()
    cursor = conn.cursor()

    schema_updates = [
        "ALTER TABLE students ADD COLUMN gender VARCHAR(10)",
        "ALTER TABLE students ADD COLUMN updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP",
        "ALTER TABLE student_records ADD COLUMN remarks TEXT",
        "ALTER TABLE student_records ADD COLUMN updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP",
        "ALTER TABLE student_records ADD INDEX idx_student_records_lrn (lrn)",
        "ALTER TABLE student_records DROP INDEX unique_student_year_grade",
        "ALTER TABLE student_records ADD UNIQUE KEY unique_student_year (lrn, school_year)",
        "ALTER TABLE student_change_logs ADD COLUMN changed_by VARCHAR(80)",
    ]

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS students (
            lrn VARCHAR(20) PRIMARY KEY,
            name VARCHAR(255),
            gender VARCHAR(10),
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        )
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS student_records (
            id INT AUTO_INCREMENT PRIMARY KEY,
            lrn VARCHAR(20),
            school_year VARCHAR(20),
            grade_level INT,
            gender VARCHAR(10),
            status VARCHAR(50),
            remarks TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY unique_student_year (lrn, school_year),
            INDEX idx_student_records_lrn (lrn),
            FOREIGN KEY (lrn) REFERENCES students(lrn)
        )
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS student_change_logs (
            id INT AUTO_INCREMENT PRIMARY KEY,
            lrn VARCHAR(20),
            field_name VARCHAR(100),
            old_value TEXT,
            new_value TEXT,
            school_year VARCHAR(20),
            grade_level INT,
            changed_by VARCHAR(80),
            changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (lrn) REFERENCES students(lrn)
        )
        """
    )

    cursor.execute(
        """
        DELETE r1
        FROM student_records r1
        JOIN student_records r2
            ON r1.lrn = r2.lrn
            AND r1.school_year = r2.school_year
            AND r1.id < r2.id
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(80) NOT NULL UNIQUE,
            password_hash VARCHAR(255) NOT NULL,
            role VARCHAR(50) NOT NULL DEFAULT 'admin',
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )

    for statement in schema_updates:
        try:
            cursor.execute(statement)
        except mysql.connector.Error as error:
            if error.errno not in (1060, 1061, 1062, 1091):
                raise

    cursor.execute(
        """
        UPDATE student_records
        SET remarks = ''
        WHERE LOWER(TRIM(remarks)) = 'nan'
        """
    )

    cursor.execute(
        """
        UPDATE student_records
        SET status = CASE
            WHEN UPPER(COALESCE(remarks, '')) LIKE '%PENDING TI%'
                OR UPPER(COALESCE(remarks, '')) LIKE '%PENDING T/I%'
                THEN 'PENDING_TRANSFER_IN'
            WHEN UPPER(COALESCE(remarks, '')) LIKE '%T/O%'
                OR UPPER(COALESCE(remarks, '')) LIKE '%T-O%'
                OR UPPER(COALESCE(remarks, '')) LIKE '%TRANSFER OUT%'
                OR UPPER(COALESCE(remarks, '')) LIKE '%TRANSFERRED OUT%'
                OR UPPER(COALESCE(remarks, '')) LIKE '%TRANSFER-OUT%'
                THEN 'TRANSFER_OUT'
            WHEN UPPER(COALESCE(remarks, '')) LIKE '%T/I%'
                OR UPPER(COALESCE(remarks, '')) LIKE '%T-I%'
                OR UPPER(COALESCE(remarks, '')) LIKE '%TRANSFER IN%'
                OR UPPER(COALESCE(remarks, '')) LIKE '%TRANSFERRED IN%'
                OR UPPER(COALESCE(remarks, '')) LIKE '%TRANSFER-IN%'
                THEN 'TRANSFER_IN'
            ELSE 'ENROLLED'
        END
        """
    )

    cursor.execute("SELECT COUNT(*) FROM users")
    user_count = cursor.fetchone()[0] or 0
    if user_count == 0:
        cursor.execute(
            """
            INSERT INTO users (username, password_hash, role)
            VALUES (%s, %s, %s)
            """,
            ("admin", generate_password_hash("admin123"), "admin"),
        )

    conn.commit()
    cursor.close()
    conn.close()
    SCHEMA_READY = True


def normalize_gender(value):
    gender = str(value).strip().upper()

    if gender.startswith("M"):
        return "MALE"
    if gender.startswith("F"):
        return "FEMALE"
    return "UNKNOWN"


def detect_status_from_remarks(remarks):
    normalized = str(remarks).upper().replace("\n", " ").strip()

    if not normalized or normalized == "NAN":
        return "ENROLLED"

    transfer_out_patterns = [
        "T/O",
        "T-O",
        "TRANSFER OUT",
        "TRANSFERRED OUT",
        "TRANSFER-OUT",
    ]

    transfer_in_patterns = [
        "T/I",
        "T-I",
        "TRANSFER IN",
        "TRANSFERRED IN",
        "TRANSFER-IN",
    ]

    if "PENDING TI" in normalized or "PENDING T/I" in normalized:
        return "PENDING_TRANSFER_IN"

    if any(pattern in normalized for pattern in transfer_out_patterns):
        return "TRANSFER_OUT"

    if any(pattern in normalized for pattern in transfer_in_patterns):
        return "TRANSFER_IN"

    return "ENROLLED"


def normalize_remarks(value):
    remarks = str(value).strip()

    if remarks.lower() == "nan":
        return ""

    return remarks


def log_change(cursor, lrn, field_name, old_value, new_value, school_year=None, grade_level=None, changed_by=None):
    old_text = "" if old_value is None else str(old_value).strip()
    new_text = "" if new_value is None else str(new_value).strip()

    if old_text == new_text:
        return False

    cursor.execute(
        """
        INSERT INTO student_change_logs
            (lrn, field_name, old_value, new_value, school_year, grade_level, changed_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """,
        (lrn, field_name, old_text, new_text, school_year, grade_level, changed_by),
    )
    return True


def upsert_student(cursor, lrn, name, gender, school_year, grade_level, changed_by=None):
    cursor.execute(
        """
        SELECT name, gender
        FROM students
        WHERE lrn = %s
        """,
        (lrn,),
    )
    existing = cursor.fetchone()

    if existing is None:
        cursor.execute(
            """
            INSERT INTO students (lrn, name, gender)
            VALUES (%s, %s, %s)
            """,
            (lrn, name, gender),
        )
        return 0

    old_name, old_gender = existing
    changes_logged = 0

    if log_change(cursor, lrn, "student.name_conflict", old_name, name, school_year, grade_level, changed_by):
        changes_logged += 1
    if log_change(cursor, lrn, "student.gender_conflict", old_gender, gender, school_year, grade_level, changed_by):
        changes_logged += 1

    return changes_logged


def upsert_student_record(cursor, lrn, school_year, grade_level, gender, status, remarks, changed_by=None):
    cursor.execute(
        """
        SELECT id, gender, status, remarks
        FROM student_records
        WHERE lrn = %s
        AND school_year = %s
        ORDER BY id DESC
        LIMIT 1
        """,
        (lrn, school_year),
    )
    existing = cursor.fetchone()

    if existing is None:
        cursor.execute(
            """
            INSERT INTO student_records (lrn, school_year, grade_level, gender, status, remarks)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (lrn, school_year, grade_level, gender, status, remarks),
        )
        return "inserted", 0

    record_id, old_gender, old_status, old_remarks = existing
    changes_logged = 0

    cursor.execute("SELECT grade_level FROM student_records WHERE id = %s", (record_id,))
    old_grade = cursor.fetchone()[0]
    if log_change(cursor, lrn, "record.grade_level", old_grade, grade_level, school_year, grade_level, changed_by):
        changes_logged += 1
    if log_change(cursor, lrn, "record.gender", old_gender, gender, school_year, grade_level, changed_by):
        changes_logged += 1
    if log_change(cursor, lrn, "record.status", old_status, status, school_year, grade_level, changed_by):
        changes_logged += 1
    if log_change(cursor, lrn, "record.remarks", old_remarks, remarks, school_year, grade_level, changed_by):
        changes_logged += 1

    cursor.execute(
        """
        UPDATE student_records
        SET grade_level = %s, gender = %s, status = %s, remarks = %s
        WHERE id = %s
        """,
        (grade_level, gender, status, remarks, record_id),
    )
    return "updated", changes_logged


def find_sf1_columns(df):
    header_row = None

    for index, row in df.head(12).iterrows():
        normalized = [str(value).upper().replace("\n", " ").strip() for value in row]

        if "LRN" in normalized and any("NAME" in value for value in normalized):
            header_row = index
            break

    if header_row is None:
        return None

    columns = {}

    for col, value in enumerate(df.iloc[header_row]):
        label = str(value).upper().replace("\n", " ").strip()

        if label == "LRN":
            columns["lrn"] = col
        elif "NAME" in label:
            columns["name"] = col
        elif "SEX" in label:
            columns["sex"] = col
        elif "REMARKS" in label:
            columns["remarks"] = col

    required = {"lrn", "name", "sex"}
    if not required.issubset(columns):
        return None

    next_row_index = header_row + 1
    data_start_row = header_row + 2

    if next_row_index < len(df):
        next_lrn = str(df.iloc[next_row_index, columns["lrn"]]).strip()
        next_lrn = re.sub(r"\.0$", "", next_lrn)
        if LRN_PATTERN.match(next_lrn):
            data_start_row = next_row_index

    columns["data_start_row"] = data_start_row
    return columns


@app.route("/upload", methods=["POST"])
@admin_required
def upload():
    pending_file = None
    pending_path = None
    try:
        ensure_schema()

        confirm_mismatch = request.form.get("confirm_metadata_mismatch") == "1"
        pending_token = request.form.get("pending_upload_token", "").strip()
        pending = None
        file = None

        if confirm_mismatch and pending_token:
            pending = pop_pending_upload(pending_token)
            if not pending:
                flash("The pending upload expired. Please select the file and upload again.", "upload_error")
                return redirect(url_for("lis_upload"))

            pending_path = pending["path"]
            pending_file = StoredUpload(pending_path, pending["filename"])
            file = pending_file
        else:
            file = request.files.get("file")

        school_year = request.form.get("school_year", "").strip()
        grade_level = request.form.get("grade_level")

        if not file or not file.filename:
            flash("No file uploaded.", "upload_error")
            return redirect(url_for("lis_upload"))

        if not is_valid_school_year(school_year):
            flash("Invalid school year. Use consecutive format YYYY-YYYY, for example 2025-2026.", "upload_error")
            return redirect(url_for("lis_upload"))

        if not grade_level or not grade_level.isdigit() or int(grade_level) not in SUPPORTED_GRADES:
            flash("Invalid grade level. Only Grades 7 to 10 are supported.", "upload_error")
            return redirect(url_for("lis_upload"))

        grade_level = int(grade_level)

        try:
            df = read_lis_upload(file)
        except ValueError as e:
            flash(str(e), "upload_error")
            return redirect(url_for("lis_upload"))

        detected_metadata = detect_upload_metadata(df)
        metadata_mismatches = get_upload_metadata_mismatches(detected_metadata, school_year, grade_level)
        if metadata_mismatches and not confirm_mismatch:
            pending_token = save_pending_upload(file)
            return render_template(
                "lis_upload.html",
                grades=SUPPORTED_GRADES,
                pending_upload={
                    "token": pending_token,
                    "filename": file.filename,
                    "school_year": school_year,
                    "grade_level": grade_level,
                    "mismatches": metadata_mismatches,
                },
            )

        columns = find_sf1_columns(df)

        print("\n===== DATA SAMPLE =====")
        print(df.head())

        if columns is None:
            flash("Column detection failed. Check the LIS/SF1 file and make sure LRN, Name, and Sex columns are present.", "upload_error")
            return redirect(url_for("lis_upload"))

        print("LRN COL:", columns["lrn"])
        print("NAME COL:", columns["name"])
        print("SEX COL:", columns["sex"])

        selected_columns = {
            columns["lrn"]: "LRN",
            columns["name"]: "NAME",
            columns["sex"]: "SEX",
        }

        if "remarks" in columns:
            selected_columns[columns["remarks"]] = "REMARKS"

        df = df.iloc[columns["data_start_row"]:]
        df = df[list(selected_columns.keys())].rename(columns=selected_columns)

        if "REMARKS" not in df.columns:
            df["REMARKS"] = ""

        df = df.dropna(subset=["LRN"])
        df["LRN"] = df["LRN"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        df = df[df["LRN"].str.match(LRN_PATTERN)]
        df = df[df["NAME"].notna()]
        df = df[df["NAME"].astype(str).str.strip() != ""]

        if df.empty:
            flash("No valid student rows found. Check that the file contains 12-digit LRNs and student names.", "upload_error")
            return redirect(url_for("lis_upload"))

        print("\n===== CLEAN DATA =====")
        print(df.head(10))

        conn = get_db_connection()
        cursor = conn.cursor()

        records_inserted = 0
        records_updated = 0
        changes_logged = 0
        changed_by = session.get("username")

        for _, row in df.iterrows():
            lrn = str(row["LRN"]).strip()
            name = str(row["NAME"]).strip()
            gender = normalize_gender(row["SEX"])
            remarks = normalize_remarks(row["REMARKS"])
            status = detect_status_from_remarks(remarks)

            print("INSERTING:", lrn, name, gender)

            changes_logged += upsert_student(cursor, lrn, name, gender, school_year, grade_level, changed_by)
            record_action, record_changes = upsert_student_record(
                cursor,
                lrn,
                school_year,
                grade_level,
                gender,
                status,
                remarks,
                changed_by,
            )
            changes_logged += record_changes

            if record_action == "inserted":
                records_inserted += 1
            else:
                records_updated += 1

        conn.commit()
        cursor.close()
        conn.close()

        flash(
            f"{records_inserted} records imported, "
            f"{records_updated} records updated, "
            f"{changes_logged} changes logged."
        )
        return redirect(url_for("lis_upload"))

    except Exception as e:
        print("\nERROR:", str(e))
        flash(f"Upload failed: {str(e)}", "upload_error")
        return redirect(url_for("lis_upload"))
    finally:
        if pending_file:
            pending_file.close()
        if pending_path:
            try:
                os.remove(pending_path)
            except OSError:
                pass


def get_db_connection():
    config = {
        "host": os.environ.get("DB_HOST", "db"),
        "port": int(os.environ.get("DB_PORT", "3306")),
        "user": os.environ.get("DB_USER", "root"),
        "password": os.environ.get("DB_PASSWORD", "root"),
        "database": os.environ.get("DB_NAME", "mydb"),
    }

    ssl_disabled = os.environ.get("DB_SSL_DISABLED", "").lower() in {"1", "true", "yes"}
    ssl_ca = os.environ.get("DB_SSL_CA")
    if not ssl_disabled and config["host"] != "db":
        config["ssl_disabled"] = False
        if ssl_ca:
            config["ssl_ca"] = ssl_ca

    return mysql.connector.connect(**config)


def get_school_years_for_computation():
    conn = get_db_connection()
    cursor = conn.cursor()
    years = get_school_years(cursor)
    cursor.close()
    conn.close()
    return years
