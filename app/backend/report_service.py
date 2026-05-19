from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def style_report_sheet(sheet):
    dark_blue = "12376F"
    light_blue = "EAF1FF"
    border_color = "D9E2F3"
    thin = Side(style="thin", color=border_color)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=dark_blue)
        cell.font = Font(color="FFFFFF", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")

    for row_number in (5, 12):
        for cell in sheet[row_number]:
            cell.fill = PatternFill("solid", fgColor=light_blue)
            cell.font = Font(bold=True, color=dark_blue)

    for column_index, column_cells in enumerate(sheet.columns, start=1):
        column = get_column_letter(column_index)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column].width = min(max(max_length + 3, 12), 42)

    sheet.freeze_panes = "A13"


def build_grade7_cohort_report_workbook(report):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cohort Report"

    sheet.merge_cells("A1:G1")
    sheet["A1"] = report["title"]
    sheet["A2"] = "Entry Cohort Start"
    sheet["B2"] = report["start_year"]
    sheet["A3"] = "Expected Grade 10 Year"
    sheet["B3"] = report["end_year"]

    summary_rows = [
        ("Summary", "Value", "", "Indicator", "Value", "", "Meaning"),
        (f"Original Grade {report['start_grade']} Entry Cohort", report["summary"]["total"], "", "On-Time Completion Rate", f"{report['rates']['on_time_completion']}%", "", "Completed Grade 10 on the expected year"),
        ("On-Time Completed", report["summary"]["on_time"], "", "Overall Completion Rate", f"{report['rates']['overall_completion']}%", "", "Completed Grade 10 even with irregular records"),
        ("Completed Grade 10", report["summary"]["completed"], "", "Irregular / Repeater", report["summary"]["delayed"], "", "Learners with repetition or irregular completion"),
        ("Transfer-Out", report["summary"]["transfer_out"], "", "For Review", report["summary"]["for_review"], "", "Records needing verification"),
        ("Incomplete", report["summary"]["incomplete"], "", "", "", "", ""),
    ]
    for offset, row in enumerate(summary_rows, start=5):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=offset, column=column, value=value)

    path_row = 12
    path_headers = ["Expected Path"] + [f"Grade {step['grade']}" for step in report["expected_path"]]
    for column, value in enumerate(path_headers, start=1):
        sheet.cell(row=path_row, column=column, value=value)
    sheet.cell(row=path_row + 1, column=1, value="School Year")
    for column, step in enumerate(report["expected_path"], start=2):
        sheet.cell(row=path_row + 1, column=column, value=step["school_year"])

    table_start = 16
    path_columns = [f"Grade {step['grade']} ({step['school_year']})" for step in report["expected_path"]]
    headers = ["LRN", "Student Name"] + path_columns + ["Final Result", "Review Note"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=table_start, column=column, value=header)

    for row_index, row_data in enumerate(report["rows"], start=table_start + 1):
        sheet.cell(row=row_index, column=1, value=row_data["lrn"])
        sheet.cell(row=row_index, column=2, value=row_data["name"])
        for column, cell in enumerate(row_data["path"], start=3):
            value = cell["status_label"]
            if cell["remarks"]:
                value = f"{value} - {cell['remarks']}"
            sheet.cell(row=row_index, column=column, value=value)
        sheet.cell(row=row_index, column=3 + len(report["expected_path"]), value=row_data["result_label"])
        sheet.cell(row=row_index, column=4 + len(report["expected_path"]), value=row_data["reason"] or "")

    if not report["rows"]:
        sheet.cell(row=table_start + 1, column=1, value=f"No Grade {report['start_grade']} learners found for this starting school year.")

    style_report_sheet(sheet)

    review_sheet = workbook.create_sheet("For Review")
    review_sheet.merge_cells("A1:F1")
    review_sheet["A1"] = "Learners for Review"
    review_headers = ["LRN", "Student Name", "Last Grade Seen", "Last School Year", "Reason", "Remarks"]
    for column, header in enumerate(review_headers, start=1):
        review_sheet.cell(row=5, column=column, value=header)

    for row_index, learner in enumerate(report["review_learners"], start=6):
        review_sheet.cell(row=row_index, column=1, value=learner["lrn"])
        review_sheet.cell(row=row_index, column=2, value=learner["name"])
        review_sheet.cell(row=row_index, column=3, value=f"Grade {learner['last_grade']}")
        review_sheet.cell(row=row_index, column=4, value=learner["last_year"])
        review_sheet.cell(row=row_index, column=5, value=learner["reason"])
        review_sheet.cell(row=row_index, column=6, value=learner["remarks"])

    if not report["review_learners"]:
        review_sheet.cell(row=6, column=1, value="No learners need review for this cohort.")

    style_report_sheet(review_sheet)

    breakdown_sheet = workbook.create_sheet("Review Breakdown")
    breakdown_sheet.merge_cells("A1:E1")
    breakdown_sheet["A1"] = "Progression Review Breakdown"
    breakdown_headers = ["Grade Level", "Transferred Out", "Irregular / Repeater", "Missing / Incomplete", "Total for Review"]
    for column, header in enumerate(breakdown_headers, start=1):
        breakdown_sheet.cell(row=5, column=column, value=header)

    for row_index, item in enumerate(report["breakdown"], start=6):
        breakdown_sheet.cell(row=row_index, column=1, value=f"Grade {item['grade']}")
        breakdown_sheet.cell(row=row_index, column=2, value=item["transferred_out"])
        breakdown_sheet.cell(row=row_index, column=3, value=item["repeated_or_delayed"])
        breakdown_sheet.cell(row=row_index, column=4, value=item["missing"])
        breakdown_sheet.cell(row=row_index, column=5, value=item["for_review"])

    style_report_sheet(breakdown_sheet)
    return workbook
